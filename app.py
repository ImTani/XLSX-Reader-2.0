import sys
import os
import webbrowser
import statistics
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import (
    FigureCanvasQTAgg
    as FigureCanvas)
from main import Ui_MainWindow
from PyQt5 import QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QSettings
from PyQt5.QtWidgets import (
    QMainWindow, QDialog, QLabel, QFileDialog,
    QApplication, QTableWidgetItem, QMessageBox,
    QSplashScreen, QVBoxLayout)


class MyApp(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.updateRecentFilesMenu()
        self.sheetDropbox.currentIndexChanged.connect(self.loadSheet)
        self.browseButton.clicked.connect(self.browseFiles)
        self.checkButton.clicked.connect(self.getInfo)
        self.saveButton.clicked.connect(self.saveFile)
        self.actionOpen_.triggered.connect(self.browseFiles)
        self.actionCheck_For_Updates.triggered.connect(self.openGithub)
        self.actionSave_2.triggered.connect(self.saveFile)
        self.actionClear_Recent_Files.triggered.connect(self.clearRecentFilesMenu)  # noqa E501
        self.searchButton.clicked.connect(self.search)
        self.searchLineEdit.returnPressed.connect(self.search)
        self.calculateButton.clicked.connect(self.calculateStatistics)

    def browseFiles(self):
        browsedFilePath = QFileDialog.getOpenFileName(
            self.centralwidget, "Open File",
            filter="Excel Files (*.xlsx)")
        self.filePath = browsedFilePath[0]
        if self.filePath:
            self.tableWidget.setRowCount(0)
            self.getSheets(self.filePath)
        else:
            self.pathLineEdit.setText("No file selected.")

    def openGithub(self):
        webbrowser.open_new_tab("https://github.com/ImTani/XLSX-Reader-2.0")

    def ReadFile(self, filePath):

        self.filePath = filePath

        wb = openpyxl.load_workbook(filePath)
        try:
            sheet = wb[self.sheet]
        except KeyError:
            return

        try:
            marks_column_index = 4
        except ValueError:
            error_dialog = QMessageBox()
            error_dialog.setIcon(QMessageBox.Critical)
            error_dialog.setWindowTitle("Error")
            error_dialog.setWindowIcon(self.icon)
            error_dialog.setText("The selected sheet does not contain the 'Marks' column.")  # noqa E501
            error_dialog.exec_()
            self.tableWidget.setRowCount(0)
            return

        mark_values_set = set()
        # Initialize a variable to store the previous row number
        prev_row_number = 0
        # Loop through all the rows and retrieve values from the "Marks" column
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True),
                                start=2):
            mark_value = row[marks_column_index - 1]
            # Skip empty cells
            if mark_value is None:
                continue
            # Skip if the current row has the same row number as the previous row # noqa E501
            if i == prev_row_number + 1:
                continue
            # Add non-empty mark values to the set
            mark_values_set.add(mark_value)
            # Update the previous row number
            prev_row_number = i
        # Convert the set to a list
        mark_values_list = list(mark_values_set)
        mark_values_list = [str(i) for i in mark_values_list]

        self.comboBox.clear()
        self.comboBox.addItems(mark_values_list)

    def addRecentFile(self, file_path):
        settings = QSettings('TaniDev', 'XL_Reader')
        recent_files = settings.value('recent_files', [])
        if file_path in recent_files:
            recent_files.remove(file_path)
        recent_files.insert(0, file_path)
        if len(recent_files) > 5:
            recent_files.pop()
        settings.setValue('recent_files', recent_files)
        self.updateRecentFilesMenu()

    def updateRecentFilesMenu(self):
        settings = QSettings('TaniDev', 'XL_Reader')
        recent_files = settings.value('recent_files', [])
        self.menuRecent_Files.clear()
        for i, file_path in enumerate(recent_files):
            action = QtWidgets.QAction(f"{i+1}. {file_path}", self)
            action.triggered.connect(lambda _,
                                     fp=file_path: self.getSheets(fp))
            self.menuRecent_Files.addAction(action)

    def clearRecentFilesMenu(self):
        self.menuRecent_Files.clear()
        settings = QSettings('TaniDev', 'XL_Reader')
        settings.setValue('recent_files', [])

    def getSheets(self, filePath):
        self.filePath = filePath

        self.sheetDropbox.clear()
        self.addRecentFile(filePath)
        self.pathLineEdit.setText(self.filePath)

        workbook = openpyxl.load_workbook(filePath)
        # Get the sheet names
        sheet_names = workbook.sheetnames
        for i in sheet_names:
            self.sheetDropbox.addItem(i)

        self.sheet = sheet_names[0]

        self.ReadFile(filePath)

    def loadSheet(self):
        self.sheet = self.sheetDropbox.currentText()
        self.comboBox.clear()
        self.maxScoreValue.setText('')
        self.minScoreValue.setText('')
        self.maxScoreValue_2.setText('')
        self.minScoreValue_2.setText('')
        self.totalStudentValue_2.setText('')
        self.totalStudentValue.setText('')
        try:
            self.ReadFile(self.filePath)
        except openpyxl.utils.exceptions.InvalidFileException:
            return

    def getInfo(self):
        # Load the workbook and active sheet
        try:
            self.subCode = int(self.comboBox.currentText())
        except ValueError:
            return
        self.organiseData()

    def organiseData(self):
        try:
            wb = openpyxl.load_workbook(window.filePath)
        except openpyxl.utils.exceptions.InvalidFileException:
            error_dialog = QMessageBox()
            error_dialog.setIcon(QMessageBox.Warning)
            error_dialog.setWindowTitle("Error")
            error_dialog.setWindowIcon(self.icon)
            error_dialog.setText("No file is selected.")  # noqa E501
            error_dialog.exec_()
            self.tableWidget.setRowCount(0)
            self.sheetDropbox.clear()
            self.comboBox.clear()
            return
        sheet = wb[self.sheet]

        max_row = sheet.max_row

        # Define the selected subject code (example: 370)
        selected_subject_code = self.subCode

        # Create a list to store the filtered data
        self.filtered_data = []

        # Loop through each row in the worksheet
        for row in range(2, max_row+1):
            # Get the value of the cell in the second column of the current row
            cell_value = sheet.cell(row=row, column=4).value

            # Check if the cell value contains the selected subject code
            if str(selected_subject_code) in str(cell_value):
                roll = sheet.cell(row=row, column=1).value

                gender = sheet.cell(row=row, column=2).value

                name = sheet.cell(row=row, column=3).value

                marks = sheet.cell(row=row+1, column=4).value

                grade = sheet.cell(row=row+1, column=5).value

                # Add the filtered data to the list
                self.filtered_data.append([roll, gender, name, marks, grade])

        self.makeTable()

    def makeTable(self):
        dataLen = len(self.filtered_data)
        self.tableWidget.setRowCount(dataLen)
        self.tableWidget.setColumnCount(5)
        self.totalStudentValue.setText(str(dataLen))
        roll = []
        gender = []
        name = []
        marks = []
        grades = []

        for i in self.filtered_data:
            roll.append(str(i[0]))
            gender.append(i[1])
            name.append(i[2])
            marks.append(str(i[3]))
            grades.append(i[4])

        for i in range(dataLen):
            self.tableWidget.setItem(i, 0, QTableWidgetItem(roll[i]))
            self.tableWidget.setItem(i, 1, QTableWidgetItem(gender[i]))
            self.tableWidget.setItem(i, 2, QTableWidgetItem(name[i]))
            self.tableWidget.setItem(i, 3, QTableWidgetItem(marks[i]))
            self.tableWidget.setItem(i, 4, QTableWidgetItem(grades[i]))

        self.tableWidget.resizeColumnsToContents()
        for column in range(self.tableWidget.columnCount()):
            current_width = self.tableWidget.columnWidth(column)
            self.tableWidget.setColumnWidth(column, current_width + 10)

        sum = 0
        for i in marks:
            try:
                sum += int(i)
            except ValueError:
                error_dialog = QMessageBox()
                error_dialog.setIcon(QMessageBox.Warning)
                error_dialog.setWindowTitle("Error")
                error_dialog.setWindowIcon(self.icon)
                error_dialog.setText(
                    "The given sheet is not a marksheet.")
                error_dialog.exec_()
                self.tableWidget.setRowCount(0)
                return
            avg = round(sum/len(marks), 2)

        self.maxScoreValue.setText(str(max(marks)))
        self.minScoreValue.setText(str(min(marks)))
        self.maxScoreValue_2.setText(name[marks.index(max(marks))])
        self.minScoreValue_2.setText(name[marks.index(min(marks))])
        self.totalStudentValue_2.setText(str(avg))

    def saveFile(self):
        # Create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # copy headers
        for column in range(self.tableWidget.columnCount()):
            header_item = (self.tableWidget.horizontalHeaderItem(column))
            if header_item is not None:
                sheet.cell(row=1, column=column+1, value=header_item.text())

        # copy data
        for row in range(self.tableWidget.rowCount()):
            for column in range(self.tableWidget.columnCount()):
                item = (self.tableWidget.item(row, column))
                if item is not None:
                    sheet.cell(row=row+2, column=column+1, value=item.text())

        savePath, _ = QFileDialog.getSaveFileName(
            self.centralwidget, "Save File",
            filter="Excel Files (*.xlsx)")

        if savePath:
            dirPath, fileNameExt = os.path.split(savePath)
            fileName, fileExt = os.path.splitext(fileNameExt)

            if fileExt != ".xlsx":
                fileExt = ".xlsx"

            savePath = os.path.join(dirPath, fileName + fileExt)

            workbook.save(savePath)
        else:
            return

    def search(self):
        search_text = self.searchLineEdit.text().lower()
        if not search_text:
            return

        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                if item is not None and search_text in item.text().lower():
                    self.tableWidget.setCurrentCell(row, col)
                    self.tableWidget.clearSelection()
                    item.setSelected(True)
                    self.tableWidget.setFocus()
                    return

        QMessageBox.warning(
            self, 'Search', f'"{search_text}" not found in the table.')

    def calculateStatistics(self):
        # Get the number of rows in the table
        row_count = self.tableWidget.rowCount()

        # Get the values in the 4th column
        column_values = []
        student_names = []
        for row_index in range(row_count):
            item = self.tableWidget.item(row_index, 3)
            student_item = self.tableWidget.item(row_index, 2)
            if item is not None and item.text() and student_item is not None and student_item.text(): # noqa E501
                column_values.append(float(item.text()))
                student_names.append(student_item.text())

        if len(column_values) == 0:
            error_dialog = QMessageBox()
            error_dialog.setIcon(QMessageBox.Warning)
            error_dialog.setWindowTitle("Error")
            error_dialog.setText(
                "The selected column does not contain any numeric values.")
            error_dialog.exec_()
            return

        # Calculate the average, median, and mode of the values
        average = sum(column_values) / len(column_values)
        median = statistics.median(column_values)
        mode = statistics.mode(column_values)

        # Format the numbers
        average = round(average, 2)
        median = round(median, 2)
        mode = round(mode, 2)

        # Create a custom dialog box
        dialog = QDialog(self)
        dialog.setWindowTitle("Statistics")
        layout = QVBoxLayout(dialog)

        # Add labels for the statistics
        average_label = QLabel(f"Average: {average}")
        median_label = QLabel(f"Median: {median}")
        mode_label = QLabel(f"Mode: {mode}")

        layout.addWidget(average_label)
        layout.addWidget(median_label)
        layout.addWidget(mode_label)

        # Create a figure and axes for the chart
        figure = plt.figure()
        axes = figure.add_subplot(111)

        # Plot the data using horizontal bar chart
        y_pos = np.arange(len(student_names))
        axes.barh(y_pos, column_values, align='center', alpha=0.5)
        axes.set_yticks(y_pos)
        axes.set_yticklabels(student_names)
        axes.set_xlabel('Marks')
        axes.set_xlim([0, 100])  # Set the x-axis limit from 0 to 100

        # Add the chart to the dialog
        chart_canvas = FigureCanvas(figure)
        layout.addWidget(chart_canvas)

        # Display the dialog
        dialog.exec_()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    pixmap = QPixmap('./resources/splashScreen.png')
    splash = QSplashScreen(pixmap)
    splash.show()
    window = MyApp()
    window.show()
    splash.finish(window)
    sys.exit(app.exec_())
